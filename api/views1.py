# views.py - Version complète avec filtres, graphiques et exports
import pandas as pd
import json
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.db.models import Q, F, ExpressionWrapper, DecimalField, IntegerField
from django.db.models.functions import Coalesce, Cast, TruncDate, TruncMonth, TruncYear
from django.db.models import Sum, Count, Avg, Max, Min
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import csv
import io

from .models import *
from .serializers1 import *

class StatisticsViewSet(viewsets.ViewSet):
    """
    ViewSet complet pour les statistiques avec filtres, graphiques et exports
    """
    
    def _apply_filters(self, queryset, filters):
        """Applique les filtres communs à tous les modèles"""
        if not filters:
            return queryset
        
        # Filtres de date
        date_filters = Q()
        if filters.get('start_date'):
            date_filters &= Q(created_at__date__gte=filters['start_date'])
        if filters.get('end_date'):
            date_filters &= Q(created_at__date__lte=filters['end_date'])
        
        if date_filters:
            queryset = queryset.filter(date_filters)
        
        # Filtres par point de vente
        if filters.get('point_of_sale'):
            if hasattr(queryset.model, 'point_of_sale'):
                queryset = queryset.filter(point_of_sale_id__in=filters['point_of_sale'])
            elif hasattr(queryset.model, 'vendor_activity'):
                queryset = queryset.filter(
                    vendor_activity__vendor__point_of_sale_id__in=filters['point_of_sale']
                )
        
        # Filtres par vendeur
        if filters.get('vendor'):
            if hasattr(queryset.model, 'vendor'):
                queryset = queryset.filter(vendor_id__in=filters['vendor'])
        
        # Filtres par région
        if filters.get('region'):
            if hasattr(queryset.model, 'region'):
                queryset = queryset.filter(region__in=filters['region'])
            elif hasattr(queryset.model, 'point_of_sale'):
                queryset = queryset.filter(point_of_sale__region__in=filters['region'])
        
        # Filtres par zone
        if filters.get('zone'):
            if hasattr(queryset.model, 'zone'):
                queryset = queryset.filter(zone__in=filters['zone'])
        
        return queryset
    
    def _get_date_range(self, period):
        """Retourne la plage de dates selon la période"""
        end_date = timezone.now()
        if period == 'today':
            start_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == 'yesterday':
            start_date = end_date - timedelta(days=1)
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=1) - timedelta(seconds=1)
        elif period == 'week':
            start_date = end_date - timedelta(days=7)
        elif period == 'month':
            start_date = end_date - timedelta(days=30)
        elif period == 'quarter':
            start_date = end_date - timedelta(days=90)
        elif period == 'year':
            start_date = end_date - timedelta(days=365)
        else:  # 30 jours par défaut
            start_date = end_date - timedelta(days=30)
        
        return start_date, end_date
    
    def _calculate_growth(self, current_value, previous_value):
        """Calcule le taux de croissance"""
        if previous_value == 0:
            return 100.0 if current_value > 0 else 0.0
        return round(((current_value - previous_value) / previous_value) * 100, 2)
    
    # ==================== DASHBOARD & RÉSUMÉ ====================
    
    @action(detail=False, methods=['get'])
    def dashboard_summary(self, request):
        """Résumé général du dashboard avec filtres"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            # Période par défaut si non spécifiée
            period = filters.get('period', 'month')
            start_date, end_date = self._get_date_range(period)
            
            # Ventes totales
            sales_qs = Sale.objects.all()
            sales_qs = self._apply_filters(sales_qs, filters)
            
            total_sales_agg = sales_qs.aggregate(
                total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
            )
            total_sales = total_sales_agg['total'] or 0
            
            # Commandes
            orders_qs = Order.objects.all()
            orders_qs = self._apply_filters(orders_qs, filters)
            total_orders = orders_qs.count()
            
            # Vendeurs ambulants actifs
            vendors_qs = MobileVendor.objects.filter(status='actif')
            if filters.get('point_of_sale'):
                vendors_qs = vendors_qs.filter(point_of_sale_id__in=filters['point_of_sale'])
            total_mobile_vendors = vendors_qs.count()
            
            # Points de vente actifs
            pos_qs = PointOfSale.objects.filter(status='actif')
            if filters.get('region'):
                pos_qs = pos_qs.filter(region__in=filters['region'])
            total_points_of_sale = pos_qs.count()
            
            # Achats actifs
            purchases_qs = Purchase.objects.all()
            purchases_qs = self._apply_filters(purchases_qs, filters)
            active_purchases = purchases_qs.count()
            
            # Calcul de la croissance
            previous_start_date = start_date - (end_date - start_date)
            previous_sales_agg = sales_qs.filter(
                created_at__gte=previous_start_date,
                created_at__lt=start_date
            ).aggregate(
                total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
            )
            previous_sales = previous_sales_agg['total'] or 0
            
            sales_growth = self._calculate_growth(float(total_sales), float(previous_sales))
            
            data = {
                'total_sales': float(total_sales),
                'total_orders': total_orders,
                'total_mobile_vendors': total_mobile_vendors,
                'total_points_of_sale': total_points_of_sale,
                'active_purchases': active_purchases,
                'sales_growth': sales_growth,
                'revenue_growth': sales_growth,  # Même calcul pour l'instant
                'period': period,
                'start_date': start_date.date().isoformat() if start_date else None,
                'end_date': end_date.date().isoformat() if end_date else None
            }
            
            return Response(data)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des statistiques: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # ==================== STATISTIQUES POINTS DE VENTE ====================
    
    @action(detail=False, methods=['get'])
    def points_of_sale_stats(self, request):
        """Statistiques par point de vente avec filtres"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            pos_stats = []
            pos_qs = PointOfSale.objects.all()
            
            if filters.get('region'):
                pos_qs = pos_qs.filter(region__in=filters['region'])
            
            for pos in pos_qs:
                # Ventes du POS
                sales_qs = Sale.objects.filter(
                    vendor_activity__vendor__point_of_sale=pos
                )
                sales_qs = self._apply_filters(sales_qs, filters)
                
                sales_agg = sales_qs.aggregate(
                    total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2)),
                    count=Count('id')
                )
                total_sales = sales_agg['total'] or 0
                
                # Commandes du POS
                orders_qs = Order.objects.filter(point_of_sale=pos)
                orders_qs = self._apply_filters(orders_qs, filters)
                total_orders = orders_qs.count()
                
                # Valeur moyenne des commandes
                avg_order_agg = orders_qs.aggregate(
                    avg=Coalesce(Avg('total'), 0, output_field=DecimalField(max_digits=10, decimal_places=2))
                )
                avg_order_value = avg_order_agg['avg'] or 0
                
                # Vendeurs ambulants
                mobile_vendors_count = pos.mobile_vendors.count()
                
                # Score de performance
                performance_score = 0
                if pos.turnover and float(pos.turnover) > 0:
                    performance_score = min(100, (float(total_sales) / float(pos.turnover)) * 100)
                
                # Calcul croissance
                period = filters.get('period', 'month')
                current_start, current_end = self._get_date_range(period)
                previous_start = current_start - (current_end - current_start)
                
                previous_sales_agg = sales_qs.filter(
                    created_at__gte=previous_start,
                    created_at__lt=current_start
                ).aggregate(
                    total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
                )
                previous_sales = previous_sales_agg['total'] or 0
                sales_growth = self._calculate_growth(float(total_sales), float(previous_sales))
                
                pos_data = {
                    'id': pos.id,
                    'name': pos.name,
                    'type': pos.type,
                    'region': pos.region,
                    'commune': pos.commune,
                    'total_sales': float(total_sales),
                    'total_orders': total_orders,
                    'average_order_value': float(avg_order_value),
                    'mobile_vendors_count': mobile_vendors_count,
                    'performance_score': round(performance_score, 2),
                    'turnover': float(pos.turnover) if pos.turnover else 0,
                    'sales_growth': sales_growth
                }
                
                pos_stats.append(pos_data)
            
            # Tri par performance
            sort_by = request.GET.get('sort_by', 'performance_score')
            reverse = request.GET.get('sort_order', 'desc') == 'desc'
            pos_stats.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
            
            return Response(pos_stats)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des stats POS: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # ==================== GRAPHIQUES ET SÉRIES TEMPORELLES ====================
    
    @action(detail=False, methods=['get'])
    def sales_chart(self, request):
        """Données pour graphique des ventes"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            group_by = filters.get('group_by', 'day')
            period = filters.get('period', 'month')
            
            start_date, end_date = self._get_date_range(period)
            
            # Agrégation par période
            if group_by == 'day':
                trunc_func = TruncDate('created_at')
                date_format = '%Y-%m-%d'
            elif group_by == 'week':
                trunc_func = TruncDate('created_at')  # Simplifié
                date_format = 'Semaine %W'
            elif group_by == 'month':
                trunc_func = TruncMonth('created_at')
                date_format = '%Y-%m'
            else:  # year
                trunc_func = TruncYear('created_at')
                date_format = '%Y'
            
            sales_qs = Sale.objects.filter(
                created_at__gte=start_date,
                created_at__lte=end_date
            )
            sales_qs = self._apply_filters(sales_qs, filters)
            
            chart_data = sales_qs.annotate(
                period=trunc_func
            ).values('period').annotate(
                sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2)),
                count=Count('id')
            ).order_by('period')
            
            labels = []
            sales_values = []
            count_values = []
            
            for data in chart_data:
                labels.append(data['period'].strftime(date_format))
                sales_values.append(float(data['sales']))
                count_values.append(data['count'])
            
            response_data = {
                'labels': labels,
                'datasets': [
                    {
                        'label': 'Ventes (FCFA)',
                        'data': sales_values,
                        'borderColor': 'rgb(75, 192, 192)',
                        'backgroundColor': 'rgba(75, 192, 192, 0.2)',
                        'yAxisID': 'y'
                    },
                    {
                        'label': 'Nombre de transactions',
                        'data': count_values,
                        'borderColor': 'rgb(255, 99, 132)',
                        'backgroundColor': 'rgba(255, 99, 132, 0.2)',
                        'yAxisID': 'y1'
                    }
                ]
            }
            
            return Response(response_data)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul du graphique: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def performance_chart(self, request):
        """Graphique de performance par vendeur/point de vente"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            chart_type = request.GET.get('chart_type', 'vendors')  # vendors, pos, products
            
            if chart_type == 'vendors':
                # Top 10 vendeurs par performance
                vendors_qs = MobileVendor.objects.filter(status='actif')
                if filters.get('point_of_sale'):
                    vendors_qs = vendors_qs.filter(point_of_sale_id__in=filters['point_of_sale'])
                
                vendors_data = []
                for vendor in vendors_qs[:10]:  # Limiter à 10 pour le graphique
                    sales_qs = Sale.objects.filter(vendor=vendor)
                    sales_qs = self._apply_filters(sales_qs, filters)
                    
                    sales_agg = sales_qs.aggregate(
                        total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
                    )
                    
                    vendors_data.append({
                        'label': vendor.full_name,
                        'value': float(sales_agg['total'] or 0),
                        'efficiency': vendor.performance or 0
                    })
                
                vendors_data.sort(key=lambda x: x['value'], reverse=True)
                
                response_data = {
                    'labels': [item['label'] for item in vendors_data],
                    'datasets': [
                        {
                            'label': 'Ventes (FCFA)',
                            'data': [item['value'] for item in vendors_data],
                            'backgroundColor': 'rgba(54, 162, 235, 0.8)'
                        }
                    ]
                }
                
            elif chart_type == 'products':
                # Top 10 produits par revenu
                products_qs = Product.objects.all()
                products_data = []
                
                for product in products_qs[:10]:
                    sales_qs = Sale.objects.filter(product_variant__product=product)
                    sales_qs = self._apply_filters(sales_qs, filters)
                    
                    sales_agg = sales_qs.aggregate(
                        revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2)),
                        quantity=Coalesce(Sum('quantity'), 0, output_field=IntegerField())
                    )
                    
                    products_data.append({
                        'label': product.name,
                        'revenue': float(sales_agg['revenue'] or 0),
                        'quantity': sales_agg['quantity'] or 0
                    })
                
                products_data.sort(key=lambda x: x['revenue'], reverse=True)
                
                response_data = {
                    'labels': [item['label'] for item in products_data],
                    'datasets': [
                        {
                            'label': 'Revenu (FCFA)',
                            'data': [item['revenue'] for item in products_data],
                            'backgroundColor': 'rgba(255, 99, 132, 0.8)'
                        },
                        {
                            'label': 'Quantité vendue',
                            'data': [item['quantity'] for item in products_data],
                            'backgroundColor': 'rgba(75, 192, 192, 0.8)'
                        }
                    ]
                }
            
            return Response(response_data)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul du graphique de performance: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # ==================== EXPORTS ====================
    
    @action(detail=False, methods=['post'])
    def export_data(self, request):
        """Export des données en CSV, Excel ou PDF"""
        try:
            serializer = ExportRequestSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            export_format = serializer.validated_data['format']
            report_type = serializer.validated_data['report_type']
            filters = serializer.validated_data.get('filters', {})
            columns = serializer.validated_data.get('columns', [])
            
            if export_format == 'csv':
                return self._export_csv(report_type, filters, columns)
            elif export_format == 'excel':
                return self._export_excel(report_type, filters, columns)
            else:
                return Response(
                    {'error': 'Format non supporté'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'export: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _export_csv(self, report_type, filters, columns):
        """Export CSV"""
        data = self._get_export_data(report_type, filters, columns)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.now().date()}.csv"'
        
        writer = csv.writer(response)
        
        if data:
            # En-têtes
            writer.writerow(data[0].keys())
            # Données
            for row in data:
                writer.writerow(row.values())
        
        return response
    
    def _export_excel(self, report_type, filters, columns):
        """Export Excel"""
        data = self._get_export_data(report_type, filters, columns)
        
        wb = Workbook()
        ws = wb.active
        ws.title = report_type
        
        if data:
            # En-têtes
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            # Données
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data.values(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Sauvegarde dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.now().date()}.xlsx"'
        
        return response
    
    def _get_export_data(self, report_type, filters, columns):
        """Récupère les données pour l'export"""
        if report_type == 'sales':
            return self._get_sales_export_data(filters, columns)
        elif report_type == 'vendors':
            return self._get_vendors_export_data(filters, columns)
        elif report_type == 'products':
            return self._get_products_export_data(filters, columns)
        elif report_type == 'pos':
            return self._get_pos_export_data(filters, columns)
        else:
            return []
    
    def _get_sales_export_data(self, filters, columns):
        """Données d'export pour les ventes"""
        sales_qs = Sale.objects.select_related(
            'product_variant__product',
            'vendor',
            'vendor_activity__vendor__point_of_sale'
        )
        sales_qs = self._apply_filters(sales_qs, filters)
        
        export_data = []
        for sale in sales_qs:
            row = {
                'Date': sale.created_at.date().isoformat(),
                'Produit': sale.product_variant.product.name if sale.product_variant and sale.product_variant.product else 'N/A',
                'Vendeur': sale.vendor.full_name if sale.vendor else 'N/A',
                'Point de vente': sale.vendor_activity.vendor.point_of_sale.name if sale.vendor_activity and sale.vendor_activity.vendor and sale.vendor_activity.vendor.point_of_sale else 'N/A',
                'Quantité': sale.quantity,
                'Prix unitaire': float(sale.total_amount / sale.quantity) if sale.quantity > 0 else 0,
                'Montant total': float(sale.total_amount),
                'Zone': getattr(sale.customer, 'zone', ''),
                'Latitude': sale.latitude or '',
                'Longitude': sale.longitude or ''
            }
            
            # Filtrer les colonnes si spécifié
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            
            export_data.append(row)
        
        return export_data
    
    def _get_vendors_export_data(self, filters, columns):
        """Données d'export pour les vendeurs"""
        vendors_qs = MobileVendor.objects.all()
        if filters.get('point_of_sale'):
            vendors_qs = vendors_qs.filter(point_of_sale_id__in=filters['point_of_sale'])
        
        export_data = []
        for vendor in vendors_qs:
            sales_qs = Sale.objects.filter(vendor=vendor)
            sales_qs = self._apply_filters(sales_qs, filters)
            
            sales_agg = sales_qs.aggregate(
                total_sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2)),
                total_quantity=Coalesce(Sum('quantity'), 0, output_field=IntegerField())
            )
            
            row = {
                'Nom complet': vendor.full_name,
                'Téléphone': vendor.phone,
                'Statut': vendor.status,
                'Type véhicule': vendor.vehicle_type,
                'Point de vente': vendor.point_of_sale.name,
                'Ventes totales': float(sales_agg['total_sales'] or 0),
                'Quantité vendue': sales_agg['total_quantity'] or 0,
                'Performance': vendor.performance or 0,
                'Date d\'inscription': vendor.date_joined.isoformat()
            }
            
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            
            export_data.append(row)
        
        return export_data
    
    def _get_products_export_data(self, filters, columns):
        """Données d'export pour les produits"""
        products_qs = Product.objects.select_related('category')
        export_data = []
        
        for product in products_qs:
            sales_qs = Sale.objects.filter(product_variant__product=product)
            sales_qs = self._apply_filters(sales_qs, filters)
            
            sales_agg = sales_qs.aggregate(
                total_revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2)),
                total_quantity=Coalesce(Sum('quantity'), 0, output_field=IntegerField())
            )
            
            total_stock = sum(variant.current_stock for variant in product.variants.all())
            
            row = {
                'Nom': product.name,
                'SKU': product.sku,
                'Catégorie': product.category.name if product.category else '',
                'Statut': product.status,
                'Revenu total': float(sales_agg['total_revenue'] or 0),
                'Quantité vendue': sales_agg['total_quantity'] or 0,
                'Stock total': total_stock,
                'Prix moyen': float(sales_agg['total_revenue'] / sales_agg['total_quantity']) if sales_agg['total_quantity'] > 0 else 0
            }
            
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            
            export_data.append(row)
        
        return export_data
    
    def _get_pos_export_data(self, filters, columns):
        """Données d'export pour les points de vente"""
        pos_qs = PointOfSale.objects.all()
        if filters.get('region'):
            pos_qs = pos_qs.filter(region__in=filters['region'])
        
        export_data = []
        for pos in pos_qs:
            sales_qs = Sale.objects.filter(vendor_activity__vendor__point_of_sale=pos)
            sales_qs = self._apply_filters(sales_qs, filters)
            
            sales_agg = sales_qs.aggregate(
                total_sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
            )
            
            orders_count = Order.objects.filter(point_of_sale=pos).count()
            vendors_count = pos.mobile_vendors.count()
            
            row = {
                'Nom': pos.name,
                'Type': pos.type,
                'Région': pos.region,
                'Commune': pos.commune,
                'Ventes totales': float(sales_agg['total_sales'] or 0),
                'Nombre de commandes': orders_count,
                'Nombre de vendeurs': vendors_count,
                'Chiffre d\'affaires': float(pos.turnover) if pos.turnover else 0,
                'Statut': pos.status
            }
            
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            
            export_data.append(row)
        
        return export_data
    
    # ==================== MÉTHODES EXISTANTES (adaptées) ====================
    
    @action(detail=False, methods=['get'])
    def mobile_vendors_stats(self, request):
        """Statistiques des vendeurs ambulants avec filtres"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            vendor_stats = []
            vendors_qs = MobileVendor.objects.all()
            
            if filters.get('point_of_sale'):
                vendors_qs = vendors_qs.filter(point_of_sale_id__in=filters['point_of_sale'])
            
            for vendor in vendors_qs:
                # Ventes
                sales_qs = Sale.objects.filter(vendor=vendor)
                sales_qs = self._apply_filters(sales_qs, filters)
                
                sales_amount_agg = sales_qs.aggregate(
                    total_sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
                )
                total_sales = sales_amount_agg['total_sales'] or 0
                
                # Achats
                purchases_qs = Purchase.objects.filter(vendor=vendor)
                purchases_qs = self._apply_filters(purchases_qs, filters)
                
                purchases_count_agg = purchases_qs.aggregate(count=Count('id'))
                purchases_amount_agg = purchases_qs.aggregate(
                    total_amount=Coalesce(Sum('amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
                )
                
                total_purchases = purchases_count_agg['count'] or 0
                total_purchase_amount = purchases_amount_agg['total_amount'] or 0
                
                # Jours d'activité
                start_date, _ = self._get_date_range(filters.get('period', 'month'))
                active_days = vendor.activities.filter(
                    timestamp__gte=start_date
                ).dates('timestamp', 'day').distinct().count()
                
                # Taux d'efficacité
                efficiency_rate = 0
                if total_purchase_amount and float(total_purchase_amount) > 0:
                    efficiency_rate = (float(total_sales) / float(total_purchase_amount)) * 100
                
                # Valeur moyenne d'achat
                average_purchase_value = 0
                if total_purchases > 0:
                    average_purchase_value = float(total_purchase_amount) / total_purchases
                
                vendor_data = {
                    'id': vendor.id,
                    'full_name': vendor.full_name,
                    'phone': vendor.phone,
                    'status': vendor.status,
                    'vehicle_type': vendor.vehicle_type,
                    'total_sales': float(total_sales),
                    'total_purchases': total_purchases,
                    'average_purchase_value': round(average_purchase_value, 2),
                    'active_days': active_days,
                    'efficiency_rate': round(efficiency_rate, 2),
                    'performance': float(vendor.performance) if vendor.performance else 0
                }
                
                vendor_stats.append(vendor_data)
            
            # Tri
            sort_by = request.GET.get('sort_by', 'efficiency_rate')
            reverse = request.GET.get('sort_order', 'desc') == 'desc'
            vendor_stats.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
            
            return Response(vendor_stats)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des stats vendeurs: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def products_stats(self, request):
        """Statistiques des produits avec filtres"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            product_stats = []
            products_qs = Product.objects.select_related('category')
            
            for product in products_qs:
                # Ventes
                sales_qs = Sale.objects.filter(product_variant__product=product)
                sales_qs = self._apply_filters(sales_qs, filters)
                
                product_revenue_agg = sales_qs.aggregate(
                    total_revenue=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=15, decimal_places=2))
                )
                product_quantity_agg = sales_qs.aggregate(
                    total_quantity=Coalesce(Sum('quantity'), 0, output_field=IntegerField())
                )
                
                total_quantity = product_quantity_agg['total_quantity'] or 0
                total_revenue = product_revenue_agg['total_revenue'] or 0
                
                # Prix moyen
                average_price = 0
                if total_quantity > 0:
                    average_price = float(total_revenue) / total_quantity
                
                # Rotation des stocks
                stock_rotation = 0
                total_stock = sum(variant.current_stock for variant in product.variants.all())
                if total_stock > 0:
                    stock_rotation = total_quantity / total_stock
                
                product_data = {
                    'id': product.id,
                    'name': product.name,
                    'sku': product.sku,
                    'category': product.category.name if product.category else '',
                    'status': product.status,
                    'total_quantity_sold': total_quantity,
                    'total_revenue': float(total_revenue),
                    'average_price': round(average_price, 2),
                    'stock_rotation': round(stock_rotation, 2)
                }
                
                product_stats.append(product_data)
            
            # Tri
            sort_by = request.GET.get('sort_by', 'total_revenue')
            reverse = request.GET.get('sort_order', 'desc') == 'desc'
            product_stats.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
            
            return Response(product_stats)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des stats produits: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    @action(detail=False, methods=['get'])
    def top_purchases_stats(self, request):
        """Statistiques des Purchases avec le plus de ventes et leur MobileVendor principal"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            # Récupérer tous les Purchases avec filtres
            purchases_qs = Purchase.objects.select_related('vendor').all()
            purchases_qs = self._apply_filters(purchases_qs, filters)
            
            purchase_stats = []
            
            for purchase in purchases_qs:
                # Calculer le total des ventes liées à ce Purchase
                sales_qs = Sale.objects.filter(customer=purchase)
                sales_qs = self._apply_filters(sales_qs, filters)
                
                sales_agg = sales_qs.aggregate(
                    total_sales_amount=Coalesce(
                        Sum('total_amount'), 
                        0, 
                        output_field=DecimalField(max_digits=15, decimal_places=2)
                    ),
                    total_sales_count=Count('id')
                )
                
                total_sales_amount = sales_agg['total_sales_amount'] or 0
                total_sales_count = sales_agg['total_sales_count'] or 0
                
                # Récupérer le MobileVendor principal (celui qui a le plus approvisionné ce Purchase)
                vendor_sales = sales_qs.values(
                    'vendor__id', 
                    'vendor__first_name', 
                    'vendor__last_name',
                    'vendor__phone'
                ).annotate(
                    vendor_total_sales=Coalesce(
                        Sum('total_amount'),
                        0,
                        output_field=DecimalField(max_digits=15, decimal_places=2)
                    ),
                    vendor_sales_count=Count('id')
                ).order_by('-vendor_total_sales')
                
                main_vendor = None
                if vendor_sales:
                    top_vendor = vendor_sales[0]
                    main_vendor = {
                        'id': top_vendor['vendor__id'],
                        'full_name': f"{top_vendor['vendor__first_name']} {top_vendor['vendor__last_name']}",
                        'phone': top_vendor['vendor__phone'],
                        'total_sales_to_purchase': float(top_vendor['vendor_total_sales'] or 0),
                        'sales_count_to_purchase': top_vendor['vendor_sales_count'] or 0
                    }
                
                # Calculer le ratio d'approvisionnement du vendeur principal
                main_vendor_ratio = 0
                if main_vendor and float(total_sales_amount) > 0:
                    main_vendor_ratio = (main_vendor['total_sales_to_purchase'] / float(total_sales_amount)) * 100
                
                purchase_data = {
                    'purchase_id': purchase.id,
                    'purchase_full_name': purchase.full_name,
                    'purchase_zone': purchase.zone,
                    'purchase_amount': float(purchase.amount),
                    'purchase_date': purchase.purchase_date.isoformat(),
                    'purchase_base': purchase.base,
                    'purchase_pushcard_type': purchase.pushcard_type,
                    'purchase_phone': purchase.phone,
                    
                    # Statistiques des ventes
                    'total_sales_amount': float(total_sales_amount),
                    'total_sales_count': total_sales_count,
                    'sales_efficiency': round((float(total_sales_amount) / float(purchase.amount)) * 100, 2) if float(purchase.amount) > 0 else 0,
                    
                    # MobileVendor principal
                    'main_mobile_vendor': main_vendor,
                    'main_vendor_ratio': round(main_vendor_ratio, 2),
                    
                    # Informations géographiques
                    'latitude': purchase.latitude,
                    'longitude': purchase.longitude
                }
                
                purchase_stats.append(purchase_data)
            
            # Trier par montant total des ventes (descendant)
            sort_by = request.GET.get('sort_by', 'total_sales_amount')
            reverse = request.GET.get('sort_order', 'desc') == 'desc'
            purchase_stats.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
            
            # Limiter aux top N si spécifié
            limit = request.GET.get('limit')
            if limit and limit.isdigit():
                purchase_stats = purchase_stats[:int(limit)]
            
            return Response(purchase_stats)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des stats des purchases: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def top_purchases_by_vendor(self, request):
        """Statistiques des Purchases groupés par MobileVendor principal"""
        try:
            serializer = FilterSerializer(data=request.GET)
            serializer.is_valid(raise_exception=True)
            filters = serializer.validated_data
            
            # Récupérer tous les MobileVendors avec leurs purchases
            vendors_qs = MobileVendor.objects.all()
            if filters.get('point_of_sale'):
                vendors_qs = vendors_qs.filter(point_of_sale_id__in=filters['point_of_sale'])
            
            vendor_purchase_stats = []
            
            for vendor in vendors_qs:
                # Récupérer les purchases de ce vendeur
                purchases_qs = Purchase.objects.filter(vendor=vendor)
                purchases_qs = self._apply_filters(purchases_qs, filters)
                
                # Calculer les statistiques agrégées pour ce vendeur
                vendor_purchases_agg = purchases_qs.aggregate(
                    total_purchase_amount=Coalesce(
                        Sum('amount'), 
                        0, 
                        output_field=DecimalField(max_digits=15, decimal_places=2)
                    ),
                    total_purchase_count=Count('id')
                )
                
                total_purchase_amount = vendor_purchases_agg['total_purchase_amount'] or 0
                total_purchase_count = vendor_purchases_agg['total_purchase_count'] or 0
                
                # Calculer les ventes totales générées par ces purchases
                # Utiliser une agrégation directe pour éviter les boucles
                purchase_ids = purchases_qs.values_list('id', flat=True)
                vendor_sales_agg = Sale.objects.filter(
                    customer_id__in=purchase_ids
                ).aggregate(
                    total_sales_amount=Coalesce(
                        Sum('total_amount'),
                        0,
                        output_field=DecimalField(max_digits=15, decimal_places=2)
                    ),
                    total_sales_count=Count('id')
                )
                
                vendor_sales_amount = vendor_sales_agg['total_sales_amount'] or 0
                vendor_sales_count = vendor_sales_agg['total_sales_count'] or 0
                
                # Calculer l'efficacité de transformation purchase -> sales
                efficiency = 0
                if float(total_purchase_amount) > 0:
                    efficiency = (float(vendor_sales_amount) / float(total_purchase_amount)) * 100
                
                # Récupérer les top purchases de ce vendeur avec leurs ventes
                top_purchases_data = []
                for purchase in purchases_qs[:5]:  # Top 5 purchases
                    purchase_sales_agg = Sale.objects.filter(customer=purchase).aggregate(
                        sales_amount=Coalesce(
                            Sum('total_amount'),
                            0,
                            output_field=DecimalField(max_digits=15, decimal_places=2)
                        )
                    )
                    
                    sales_amount = purchase_sales_agg['sales_amount'] or 0
                    purchase_efficiency = 0
                    if float(purchase.amount) > 0:
                        purchase_efficiency = (float(sales_amount) / float(purchase.amount)) * 100
                    
                    top_purchases_data.append({
                        'id': purchase.id,
                        'full_name': purchase.full_name,
                        'zone': purchase.zone,
                        'purchase_amount': float(purchase.amount),
                        'sales_amount': float(sales_amount),
                        'efficiency': round(purchase_efficiency, 2)
                    })
                
                # Trier les top purchases par montant des ventes
                top_purchases_data.sort(key=lambda x: x['sales_amount'], reverse=True)
                
                vendor_data = {
                    'vendor_id': vendor.id,
                    'vendor_full_name': vendor.full_name,
                    'vendor_phone': vendor.phone,
                    'vendor_status': vendor.status,
                    'point_of_sale': vendor.point_of_sale.name,
                    
                    # Statistiques des purchases
                    'total_purchase_amount': float(total_purchase_amount),
                    'total_purchase_count': total_purchase_count,
                    'average_purchase_value': round(float(total_purchase_amount) / total_purchase_count, 2) if total_purchase_count > 0 else 0,
                    
                    # Statistiques des ventes générées
                    'total_sales_from_purchases': float(vendor_sales_amount),
                    'total_sales_count': vendor_sales_count,
                    'purchase_to_sales_efficiency': round(efficiency, 2),
                    
                    # Top purchases
                    'top_purchases': top_purchases_data
                }
                
                vendor_purchase_stats.append(vendor_data)
            
            # Trier par efficacité ou montant total des ventes
            sort_by = request.GET.get('sort_by', 'total_sales_from_purchases')
            reverse = request.GET.get('sort_order', 'desc') == 'desc'
            vendor_purchase_stats.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
            
            return Response(vendor_purchase_stats)
            
        except Exception as e:
            return Response(
                {'error': f'Erreur lors du calcul des stats purchases par vendeur: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )